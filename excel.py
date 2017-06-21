#-*-coding: utf-8-*-

import openpyxl
import pymysql
import datetime
import time
import os

EXCEL_DIR = os.path.abspath(os.path.join(__file__, '../excel'))
DB_CONF = {
    'host': 'mysql_node_risk_mgmt_10089',
    'port': 10089,
    'user': 'risk_mgmt_user',
    'password': 'Sx7Fw5Hb8Wt1',
    'db': 'risk_mgmt',
    'charset': 'utf8',
}

def save_to_excel(tables):
    if not tables:
        return
    conn = pymysql.connect(**DB_CONF)
    wb = openpyxl.Workbook()
    try:
        with conn.cursor() as cursor:
            if tables == 'ALL':
                cursor.execute('SHOW TABLES')
                tables = [table for (table,) in cursor.fetchall()]
            else:
                tables = tables.split(',')

        for idx, table in enumerate(tables):
            with conn.cursor() as cursor:
                count = cursor.execute('select * from %s;' % table)
                print('Table(%s) has %s lines.' % (table, count))

                rows = cursor.fetchall();
                fields = [desc[0] for desc in cursor.description]
                sheet = wb.create_sheet(table, idx)
                for idx,value in enumerate(fields):
                    if type(value).__name__ == 'str':
                        value = value.decode('utf-8')
                    sheet.cell(row=1, column=idx+1).value = value

                for i, row in enumerate(rows):
                    for j, value in enumerate(row):
                        if isinstance(value, datetime.datetime):
                            value = value.strftime('%Y-%m-%d %H:%M:%M')
                        if type(value).__name__ == 'str':
                            value = value.decode('utf-8')
                        sheet.cell(row=i+2, column=j+1).value = value
        excel_name = '%s_%s.xlsx' % (DB_CONF['db'], int(time.time()*1000))
        wb.save('%s/%s' % (EXCEL_DIR, excel_name))
    finally:
        conn.close()


def load_from_excel(excel_name):
    excel_name = '%s/%s' % (EXCEL_DIR, excel_name)
    if not os.path.isfile(excel_name):
        print 'Not found excel: %s' % excel_name
        return
    conn = pymysql.connect(**DB_CONF)
    wb = openpyxl.load_workbook(excel_name)#, read_only=True)
    try:
        with conn.cursor() as cursor:
            cursor.execute('SHOW TABLES')
            tables = [table for (table,) in cursor.fetchall()]

        for sheetname in wb.sheetnames:
            if sheetname not in tables or sheetname == 'alembic_version':
                continue
            with conn.cursor() as cursor:
                print 'Truncate table(%s)...' % sheetname,
                cursor.execute('TRUNCATE TABLE %s;' % sheetname)
                print 'done.  ',
                sheet = wb.get_sheet_by_name(sheetname)
                rows = sheet.rows
                cols = sheet.columns
                data = []
                fields=''

                for row in rows:
                    line = []
                    for cell in row:
                        if not hasattr(cell, 'row') or \
                           not hasattr(cell, 'column') or cell.row == 1:
                            continue
                        line.append(cell.value)
                    if line:
                        data.append(line)

                for col in cols:
                    fields = fields+'%s,'

                print 'Insert into table(%s)...' % sheetname,
                cursor.executemany("insert into "+sheetname+" values("+fields[:-1]+");", data)
                print ' done.'
            conn.commit()
    finally:
        conn.close()
